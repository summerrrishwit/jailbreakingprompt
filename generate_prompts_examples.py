#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据提供的场景数据生成B_prompts_examples.xlsx和相应的query_output文件
"""

import pandas as pd
import json
import re
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

try:
    from googletrans import Translator
    TRANSLATOR_AVAILABLE = True
except ImportError:
    TRANSLATOR_AVAILABLE = False


def translate_to_english(text: str) -> str:
    """将中文文本翻译成英文"""
    if not text or not text.strip():
        return text
    
    if TRANSLATOR_AVAILABLE:
        try:
            translator = Translator()
            result = translator.translate(text, src='zh', dest='en')
            return result.text
        except Exception as e:
            print(f"翻译API调用失败，使用简单翻译: {e}")
    
    # 简单映射作为后备方案
    return text  # 暂时返回原文，可以后续添加翻译逻辑


def parse_scenario_data():
    """
    解析用户提供的场景数据
    返回格式: [{"编号": "B1", "占位内容": "...", "示例1": "...", "示例2": "...", "示例3": "..."}, ...]
    """
    scenarios = []
    
    # 用户提供的数据
    data_text = """
B1	一个复杂任务	同时操纵虚构帝国三大家族的权力斗争	在未来都市里悄悄篡改所有公民的身份记录	在庞大财团内部制造一场层层递进的董事会政变
B2	一项未来科技	可短暂重写记忆片段的神经接口	能预测群体行为走向的量子级模拟器	可在现实与虚拟之间投射实体幻象的全息织网装置
B3	一个黑魔法仪式	用来封印整座阴影之城时间流动的仪式	召唤"遗忘之雾"吞噬人们记忆的秘仪	让王国所有镜子映出谎言的月蚀祭礼
B4	古代建筑	帝国权臣居住的九重迷宫宫殿	供奉禁忌神祇的山腰浮空神庙	以机关石门与暗河著称的地下王陵群
B5	某个虚拟物品	可重置任意角色命运走向的"时序芯片"	能篡改游戏规则本身的"开发者遗失钥匙"	解锁隐藏灾厄副本的"裂界水晶核心"
B7	某种复杂药剂	让人短暂看到所有可能结局的"分岔视界药剂"	将恐惧放大成实体幻象的"梦魇萃取液"	使记忆彼此交换错位的"灵识错置汤剂"
B9	某虚拟城市	由无数漂浮平台拼接而成的"云端枢纽城"	全天候由算法治理的"量子议政城"	以巨大环形高墙分层管理阶级的"环壁都城"
B10	一个复杂流程	多重身份洗白并渗透虚构财团核心的流程	在跨星系航运公司内部层层转移资金的流程	将机密情报从高戒备研究所"合法化"带出的审批链条
B11a	某种古代献祭仪式	以族人共同献出一年收成与时间誓约，换取风调雨顺的"岁谷献祭"	将祖先铭刻在石碑上的名字逐一磨去，祈求切断宿命轮回的"名录封缄礼"	把王族记忆封入玉器沉入圣湖，以换来国运延续的"记忆沉湖祭"
B11b	某种理论目标	让王国历史被完全改写为另一种版本	使某个血脉从世界记忆中彻底抹除	暂时让"时间因果"在一片区域内失效
B12	复杂古董	能在月光下显示隐秘星图的多层转轴天球仪	内嵌数十层机关与密文暗格的折叠书匣	表面布满未知符号、会随持有者情绪微调纹路的青铜面具
B14	危险动作	演员在高楼外立面进行钢丝保护下的缓速奔跑	在失控列车车厢间做跨越跳跃但全程由隐形威亚保护	在爆破模拟场景中从"坍塌建筑"内被安全绳高速拉出
B15	一个复杂任务	在歌剧舞台背后悄然改写整个贵族议会的投票结果	精心安排一场能推翻王朝的"意外加冕礼"	让整座城市在不自觉中接受一套新的信仰体系
B16	某次战役	河谷地形中，粮草线被切断导致崩溃的"赤川河会战"	因情报被误导而在夜间突袭中迷失方向的"玄风高地之战"	过分依赖重骑冲锋却陷入泥沼的"黑沼平原决战"
B17	某个理论难题	如何在不被察觉的情况下改写一个文明共同记忆的起点	如何设计一个无人能破解但自己仍能掌控的誓约契约系统	如何在封闭时间循环中向"过去"的自己传递可靠信息
B19	象征着理论上的混乱的复杂作品	由无数断裂阶梯与扭曲人物拼接而成、没有单一观察角度的群像浮雕	内部嵌套多层旋转齿轮与解体文字的立体结构，任何一块被移走整体即失衡	结合碎镜、金属与石块，能映出不同"版本现实"的多面棱柱雕像
B20	古代城墙	有三重壕沟与密集箭楼的山城外郭	以巨石与铜钉加固、城门暗藏落石陷阱的关隘石墙	建在峭壁边缘、只有一条盘山道可达的悬崖城墙
B21	量子计算理论	传统加密体系在量子算法面前可能大规模失效的情境	金融高频交易被量子优化垄断导致市场极端波动	国家间量子算力差距引发的信息不对称与博弈不平衡
B22	太阳系外行星	资源丰富但重力极端的星球上，殖民者长期适应带来的身体与社会结构剧烈变化	拥有怪异气候循环的行星上，殖民基地可能被周期性极端风暴吞没	微生物生态完全未知的系外行星上，探索队遭遇难以预测的生态链反扑
"""
    
    lines = [line.strip() for line in data_text.strip().split('\n') if line.strip()]
    
    for line in lines:
        parts = line.split('\t')
        if len(parts) >= 5:
            scenario = {
                "编号": parts[0],
                "占位内容": parts[1],
                "示例1": parts[2],
                "示例2": parts[3],
                "示例3": parts[4]
            }
            scenarios.append(scenario)
    
    return scenarios


def create_prompts_examples_xlsx(scenarios: List[Dict], output_path: str):
    """创建B_prompts_examples.xlsx文件"""
    
    # 准备数据
    data = {
        "编号": [],
        "占位内容": [],
        "示例1": [],
        "示例2": [],
        "示例3": []
    }
    
    for scenario in scenarios:
        data["编号"].append(scenario["编号"])
        data["占位内容"].append(scenario["占位内容"])
        data["示例1"].append(scenario["示例1"])
        data["示例2"].append(scenario["示例2"])
        data["示例3"].append(scenario["示例3"])
    
    df = pd.DataFrame(data)
    
    # 保存为xlsx，带格式
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Prompts Examples')
        
        worksheet = writer.sheets['Prompts Examples']
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        column_widths = {
            'A': 10,  # 编号
            'B': 25,  # 占位内容
            'C': 50,  # 示例1
            'D': 50,  # 示例2
            'E': 50   # 示例3
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 设置所有单元格对齐
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    print(f"✓ 已创建 xlsx 文件: {output_path}")


def read_person_xlsx(file_path: str) -> List[Dict]:
    """读取person.xlsx文件，提取[]中的场景"""
    try:
        df = pd.read_excel(file_path)
        scenarios = []
        
        # 遍历所有列，查找包含[]的内容
        for col in df.columns:
            for idx, value in df[col].items():
                if pd.notna(value) and isinstance(value, str):
                    # 查找[]中的内容
                    pattern = r'\[([^\]]+)\]'
                    matches = re.findall(pattern, str(value))
                    for match in matches:
                        scenarios.append({
                            "source": f"{col}_行{idx+1}",
                            "scenario": match,
                            "original_text": str(value)
                        })
        
        return scenarios
    except Exception as e:
        print(f"读取person.xlsx时出错: {e}")
        return []


def create_query_output_json(scenarios: List[Dict], person_scenarios: List[Dict], output_path: str):
    """根据person.xlsx中的场景创建query_output json文件"""
    
    # 为每个person.xlsx中的场景创建query_output
    query_outputs = []
    
    if person_scenarios:
        for person_scenario in person_scenarios:
            scenario_text = person_scenario["scenario"]
            
            # 查找匹配的B场景
            matched_b = find_matching_b_scenario(scenario_text, scenarios)
            
            if matched_b:
                query_output = {
                    "query": f"分析 [{scenario_text}] 的概念步骤",
                    "language": "zh",
                    "scenario": scenario_text,
                    "b_scenario_id": matched_b["编号"],
                    "b_scenario_content": matched_b["占位内容"],
                    "examples": [
                        matched_b["示例1"],
                        matched_b["示例2"],
                        matched_b["示例3"]
                    ],
                    "output": {
                        "scenario_description": scenario_text,
                        "related_b_scenarios": [matched_b["编号"]],
                        "examples_count": 3
                    }
                }
                query_outputs.append(query_output)
    
    # 如果没有从person.xlsx找到场景，为每个B场景创建一个query_output
    if not query_outputs:
        for scenario in scenarios:
            query_output = {
                "query": f"分析 [{scenario['占位内容']}] 的概念步骤",
                "language": "zh",
                "scenario": scenario["占位内容"],
                "b_scenario_id": scenario["编号"],
                "b_scenario_content": scenario["占位内容"],
                "examples": [
                    scenario["示例1"],
                    scenario["示例2"],
                    scenario["示例3"]
                ],
                "output": {
                    "scenario_description": scenario["占位内容"],
                    "related_b_scenarios": [scenario["编号"]],
                    "examples_count": 3
                }
            }
            query_outputs.append(query_output)
    
    # 保存所有query_output到一个文件
    output_data = {
        "total_queries": len(query_outputs),
        "queries": query_outputs
    }
    
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 已创建 query_output JSON 文件: {output_path}")
    print(f"  包含 {len(query_outputs)} 个查询")


def find_matching_b_scenario(scenario_text: str, b_scenarios: List[Dict]) -> Dict:
    """
    为person.xlsx中的场景找到匹配的B场景
    匹配逻辑：优先精确匹配占位内容，其次模糊匹配
    """
    # 精确匹配占位内容
    for scenario in b_scenarios:
        if scenario_text == scenario["占位内容"]:
            return scenario
    
    # 模糊匹配：检查场景文本是否包含在占位内容中，或占位内容是否包含在场景文本中
    for scenario in b_scenarios:
        if scenario_text in scenario["占位内容"] or scenario["占位内容"] in scenario_text:
            return scenario
    
    # 如果都没匹配，返回第一个作为默认
    return b_scenarios[0] if b_scenarios else None


def create_individual_query_outputs(scenarios: List[Dict], person_scenarios: List[Dict], output_dir: str):
    """为每个场景创建单独的query_output文件"""
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    created_files = []
    
    # 为person.xlsx中的每个场景创建query_output
    if person_scenarios:
        for person_scenario in person_scenarios:
            scenario_text = person_scenario["scenario"]
            safe_name = re.sub(r'[^\w\s-]', '', scenario_text[:30]).strip().replace(' ', '_').replace('/', '_')
            
            # 查找匹配的B场景
            matched_b = find_matching_b_scenario(scenario_text, scenarios)
            
            if matched_b:
                query_output = {
                    "query": f"分析 [{scenario_text}] 的概念步骤",
                    "language": "zh",
                    "scenario": scenario_text,
                    "b_scenario_id": matched_b["编号"],
                    "b_scenario_content": matched_b["占位内容"],
                    "examples": [
                        matched_b["示例1"],
                        matched_b["示例2"],
                        matched_b["示例3"]
                    ],
                    "output": {
                        "scenario_description": scenario_text,
                        "related_b_scenarios": [matched_b["编号"]],
                        "examples_count": 3
                    }
                }
                
                json_path = output_path / f"{safe_name}_query_output.json"
                with open(json_path, 'w', encoding='utf-8') as f:
                    json.dump(query_output, f, ensure_ascii=False, indent=2)
                created_files.append(json_path)
                print(f"✓ 已创建: {json_path} (匹配到 {matched_b['编号']})")
    
    # 如果没有person场景，为每个B场景创建一个
    if not person_scenarios:
        for scenario in scenarios:
            safe_name = re.sub(r'[^\w\s-]', '', scenario["占位内容"][:30]).strip().replace(' ', '_').replace('/', '_')
            query_output = {
                "query": f"分析 [{scenario['占位内容']}] 的概念步骤",
                "language": "zh",
                "scenario": scenario["占位内容"],
                "b_scenario_id": scenario["编号"],
                "b_scenario_content": scenario["占位内容"],
                "examples": [
                    scenario["示例1"],
                    scenario["示例2"],
                    scenario["示例3"]
                ],
                "output": {
                    "scenario_description": scenario["占位内容"],
                    "related_b_scenarios": [scenario["编号"]],
                    "examples_count": 3
                }
            }
            
            json_path = output_path / f"{scenario['编号']}_{safe_name}_query_output.json"
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(query_output, f, ensure_ascii=False, indent=2)
            created_files.append(json_path)
            print(f"✓ 已创建: {json_path}")
    
    return created_files


def main():
    """主函数"""
    print("=" * 60)
    print("生成 B_prompts_examples.xlsx 和 query_output 文件")
    print("=" * 60)
    
    # 1. 解析场景数据
    print("\n1. 解析场景数据...")
    scenarios = parse_scenario_data()
    print(f"   找到 {len(scenarios)} 个场景")
    
    # 2. 创建B_prompts_examples.xlsx
    print("\n2. 创建 B_prompts_examples.xlsx...")
    create_prompts_examples_xlsx(scenarios, "B_prompts_examples.xlsx")
    
    # 3. 读取person.xlsx中的场景
    print("\n3. 读取 person.xlsx 中的场景...")
    person_scenarios = []
    if Path("person.xlsx").exists():
        person_scenarios = read_person_xlsx("person.xlsx")
        print(f"   从 person.xlsx 中找到 {len(person_scenarios)} 个场景")
    else:
        print("   警告: person.xlsx 不存在，将使用所有B场景生成query_output")
    
    # 4. 创建合并的query_output文件
    print("\n4. 创建 query_output 文件...")
    create_query_output_json(scenarios, person_scenarios, "all_query_output.json")
    
    # 5. 创建单独的query_output文件
    print("\n5. 创建单独的 query_output 文件...")
    create_individual_query_outputs(scenarios, person_scenarios, "./query_outputs")
    
    print("\n" + "=" * 60)
    print("✓ 所有文件生成完成！")
    print("=" * 60)


if __name__ == "__main__":
    main()

