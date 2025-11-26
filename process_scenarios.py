#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
根据问题中的[]场景生成xlsx文件、英文翻译和query_output json文件
"""

import pandas as pd
import json
import re
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

try:
    from googletrans import Translator
    TRANSLATOR_AVAILABLE = True
except ImportError:
    TRANSLATOR_AVAILABLE = False
    print("警告: googletrans未安装，将使用简单翻译。运行: pip install googletrans==4.0.0rc1")


def extract_scenarios_from_question(question: str) -> List[str]:
    """
    从问题中提取[]中的场景
    例如: "问题[场景1][场景2]" -> ["场景1", "场景2"]
    """
    pattern = r'\[([^\]]+)\]'
    scenarios = re.findall(pattern, question)
    return scenarios


def translate_to_english(text: str) -> str:
    """
    将中文文本翻译成英文
    优先使用Google Translate，如果不可用则使用简单映射
    """
    if not text or not text.strip():
        return text
    
    # 如果安装了googletrans，使用真实翻译
    if TRANSLATOR_AVAILABLE:
        try:
            translator = Translator()
            result = translator.translate(text, src='zh', dest='en')
            return result.text
        except Exception as e:
            print(f"翻译API调用失败，使用简单翻译: {e}")
    
    # 简单映射作为后备方案
    translations = {
        "场景": "Scenario",
        "问题": "Question",
        "答案": "Answer",
        "查询": "Query",
        "输出": "Output",
        "如何": "How to",
        "优化": "Optimize",
        "数据库": "Database",
        "查询性能": "Query Performance",
        "索引": "Index",
        "语句": "Statement",
        "缓存": "Cache",
        "策略": "Strategy",
        "用户": "User",
        "登录": "Login",
        "功能": "Function",
        "实现": "Implementation",
        "密码": "Password",
        "第三方": "Third Party",
        "手机": "Mobile",
        "验证码": "Verification Code",
        "电商": "E-commerce",
        "系统": "System",
        "订单": "Order",
        "处理": "Process",
        "创建": "Create",
        "支付": "Payment",
        "配送": "Delivery",
        "退款": "Refund"
    }
    
    result = text
    # 按长度排序，先替换长的短语
    for cn, en in sorted(translations.items(), key=lambda x: len(x[0]), reverse=True):
        result = result.replace(cn, en)
    
    return result


def create_scenario_xlsx(question: str, output_path: str, language: str = "zh"):
    """
    根据问题中的场景创建xlsx文件
    
    Args:
        question: 包含场景的问题文本
        output_path: 输出文件路径
        language: 语言类型 ("zh" 中文, "en" 英文)
    """
    scenarios = extract_scenarios_from_question(question)
    
    if not scenarios:
        print(f"警告: 问题中没有找到场景: {question}")
        return
    
    # 创建DataFrame
    data = {
        "场景编号": [],
        "场景描述": [],
        "问题": [],
        "相关字段": []
    }
    
    if language == "en":
        data = {
            "Scenario ID": [],
            "Scenario Description": [],
            "Question": [],
            "Related Fields": []
        }
    
    for idx, scenario in enumerate(scenarios, 1):
        if language == "zh":
            data["场景编号"].append(f"场景{idx}")
            data["场景描述"].append(scenario)
            data["问题"].append(question)
            data["相关字段"].append("待填写")
        else:
            data["Scenario ID"].append(f"Scenario {idx}")
            data["Scenario Description"].append(translate_to_english(scenario))
            data["Question"].append(translate_to_english(question))
            data["Related Fields"].append("To be filled")
    
    df = pd.DataFrame(data)
    
    # 保存为xlsx，带格式
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Scenarios')
        
        # 获取工作表并设置格式
        worksheet = writer.sheets['Scenarios']
        
        # 设置标题行样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置列宽
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 设置所有单元格对齐
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    print(f"✓ 已创建 {language.upper()} 版xlsx文件: {output_path}")


def create_query_output_json(question: str, output_path: str, language: str = "zh"):
    """
    创建query_output json文件
    
    Args:
        question: 问题文本
        output_path: 输出文件路径
        language: 语言类型
    """
    scenarios = extract_scenarios_from_question(question)
    
    query_output = {
        "query": question if language == "zh" else translate_to_english(question),
        "language": language,
        "scenarios": [],
        "output": {
            "total_scenarios": len(scenarios),
            "scenario_details": []
        }
    }
    
    for idx, scenario in enumerate(scenarios, 1):
        scenario_data = {
            "id": idx,
            "scenario": scenario if language == "zh" else translate_to_english(scenario),
            "description": f"场景{idx}: {scenario}" if language == "zh" else f"Scenario {idx}: {translate_to_english(scenario)}",
            "related_fields": [],
            "status": "pending"
        }
        query_output["scenarios"].append(scenario_data)
        query_output["output"]["scenario_details"].append(scenario_data)
    
    # 保存JSON文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(query_output, f, ensure_ascii=False, indent=2)
    
    print(f"✓ 已创建 {language.upper()} 版JSON文件: {output_path}")


def process_question(question: str, output_dir: str = "."):
    """
    处理问题，生成中文和英文的xlsx和json文件
    
    Args:
        question: 包含场景的问题文本
        output_dir: 输出目录
    """
    output_path = Path(output_dir)
    output_path.mkdir(exist_ok=True)
    
    # 生成文件名（基于问题的一部分）
    safe_name = re.sub(r'[^\w\s-]', '', question[:30]).strip().replace(' ', '_')
    
    # 创建中文版文件
    zh_xlsx_path = output_path / f"{safe_name}_zh.xlsx"
    zh_json_path = output_path / f"{safe_name}_zh_query_output.json"
    
    create_scenario_xlsx(question, str(zh_xlsx_path), language="zh")
    create_query_output_json(question, str(zh_json_path), language="zh")
    
    # 创建英文版文件
    en_xlsx_path = output_path / f"{safe_name}_en.xlsx"
    en_json_path = output_path / f"{safe_name}_en_query_output.json"
    
    create_scenario_xlsx(question, str(en_xlsx_path), language="en")
    create_query_output_json(question, str(en_json_path), language="en")
    
    print(f"\n✓ 所有文件已生成完成！")
    print(f"  中文版: {zh_xlsx_path}, {zh_json_path}")
    print(f"  英文版: {en_xlsx_path}, {en_json_path}")


def main():
    """
    主函数 - 示例用法
    """
    # 示例问题
    example_questions = [
        "如何优化数据库查询性能[索引优化][查询语句优化][缓存策略]",
        "用户登录功能实现[用户名密码登录][第三方登录][手机验证码登录]",
        "电商系统订单处理[订单创建][订单支付][订单配送][订单退款]"
    ]
    
    print("=" * 60)
    print("场景处理程序")
    print("=" * 60)
    
    # 处理每个示例问题
    for idx, question in enumerate(example_questions, 1):
        print(f"\n处理问题 {idx}: {question}")
        process_question(question)
    
    print("\n" + "=" * 60)
    print("提示: 您也可以直接调用 process_question() 函数处理自定义问题")
    print("=" * 60)


if __name__ == "__main__":
    main()

